# modules/excel_payroll.py
# Requires: pip install openpyxl
import io
from openpyxl import Workbook
from openpyxl.styles import (Font, PatternFill, Alignment, Border, Side,
                              numbers)
from openpyxl.utils import get_column_letter

BLUE      = "185FA5"
BLUE_LITE = "E6F1FB"
GRAY      = "F5F4F0"
RED_LITE  = "FCEBEB"
GREEN_LITE= "EAF3DE"

def _border(style="thin"):
    s = Side(style=style, color="D5D3CB")
    return Border(left=s, right=s, top=s, bottom=s)

def _fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def _hdr(ws, row, col, value, bold=True, bg=BLUE, fg="FFFFFF", wrap=False):
    c = ws.cell(row=row, column=col, value=value)
    c.font      = Font(bold=bold, color=fg, size=9)
    c.fill      = _fill(bg)
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=wrap)
    c.border    = _border()
    return c

def _cell(ws, row, col, value, fmt=None, bold=False, bg=None, align="left"):
    c = ws.cell(row=row, column=col, value=value)
    c.font      = Font(bold=bold, size=9)
    c.alignment = Alignment(horizontal=align, vertical="center")
    c.border    = _border()
    if bg:
        c.fill = _fill(bg)
    if fmt:
        c.number_format = fmt
    return c

def generate_payroll_register(employees, calc_salary_fn, year, month, month_name):
    wb = Workbook()

    # ── Sheet 1: Payroll Register ──────────────────────────────────────────
    ws = wb.active
    ws.title = f"Payroll {month_name} {year}"
    ws.freeze_panes = "B4"
    ws.row_dimensions[1].height = 28
    ws.row_dimensions[2].height = 16
    ws.row_dimensions[3].height = 30

    # Title
    ws.merge_cells("A1:P1")
    title = ws["A1"]
    title.value     = f"PAYROLL REGISTER — {month_name.upper()} {year}"
    title.font      = Font(bold=True, size=13, color=BLUE)
    title.alignment = Alignment(horizontal="center", vertical="center")
    title.fill      = _fill(BLUE_LITE)

    # Section headers row 2
    for col, (label, span_end) in enumerate([
        ("EMPLOYEE DETAILS", 4),
        ("EARNINGS", 8),
        ("DEDUCTIONS", 13),
        ("NET", 14),
    ], start=1):
        pass  # handled below via merge

    merges = [(1,4,"EMPLOYEE DETAILS"), (5,9,"EARNINGS"),
              (10,14,"DEDUCTIONS"), (15,15,"NET PAY")]
    col = 1
    for s, e, label in merges:
        ws.merge_cells(start_row=2, start_column=s, end_row=2, end_column=e)
        _hdr(ws, 2, s, label, bg=BLUE if label!="NET PAY" else "0C447C")

    # Column headers row 3
    headers = [
        "Sr.", "Name", "Department", "Designation",
        "Gross (₹)", "Basic (₹)", "HRA (₹)", "Special (₹)", "Days Worked",
        "PF (₹)", "ESI (₹)", "Prof. Tax (₹)", "TDS (₹)", "Total Ded. (₹)",
        "Net Pay (₹)"
    ]
    col_widths = [5, 22, 16, 18, 12, 12, 10, 10, 10, 10, 8, 10, 10, 12, 13]
    for i, (h, w) in enumerate(zip(headers, col_widths), start=1):
        _hdr(ws, 3, i, h, bg="1A3A5C", wrap=True)
        ws.column_dimensions[get_column_letter(i)].width = w

    # Data rows
    totals = [0] * 15
    INR = '#,##0.00'
    for sr, emp in enumerate(employees, start=1):
        slip = calc_salary_fn(emp, year, month)
        row  = 3 + sr
        bg   = GRAY if sr % 2 == 0 else "FFFFFF"

        values = [
            sr, emp["name"], emp["dept"], emp["role"],
            slip["gross"], slip["basic"], slip["hra"], slip["special"],
            slip.get("worked_days", "—"),
            slip["pf"] if emp["pf"] else 0,
            slip["esi"] if emp["esi"] else 0,
            slip["pt"], slip["tds"], slip["total_ded"],
            slip["net"]
        ]
        fmts   = [None, None, None, None, INR, INR, INR, INR, None,
                  INR, INR, INR, INR, INR, INR]
        aligns = ["center","left","left","left",
                  "right","right","right","right","center",
                  "right","right","right","right","right","right"]

        for col, (val, fmt, aln) in enumerate(zip(values, fmts, aligns), start=1):
            _cell(ws, row, col, val, fmt=fmt, bg=bg, align=aln)
            if isinstance(val, (int, float)):
                totals[col-1] = totals[col-1] + val

    # Totals row
    trow = 3 + len(employees) + 1
    ws.merge_cells(start_row=trow, start_column=1, end_row=trow, end_column=4)
    c = ws.cell(row=trow, column=1, value="TOTALS")
    c.font = Font(bold=True, size=9, color="FFFFFF")
    c.fill = _fill(BLUE)
    c.alignment = Alignment(horizontal="center", vertical="center")
    c.border = _border()

    skip = {0,1,2,3,8}   # Sr., Name, Dept, Role, Days Worked — no total
    for col in range(5, 16):
        val = totals[col-1] if col-1 not in skip else ""
        c = ws.cell(row=trow, column=col, value=val if val else "")
        c.font      = Font(bold=True, size=9, color="FFFFFF")
        c.fill      = _fill(BLUE)
        c.alignment = Alignment(horizontal="right", vertical="center")
        c.border    = _border()
        if val:
            c.number_format = INR

    ws.row_dimensions[trow].height = 18

    # ── Sheet 2: PF & ESI Register ────────────────────────────────────────
    ws2 = wb.create_sheet("PF & ESI Register")
    ws2.freeze_panes = "A3"

    ws2.merge_cells("A1:H1")
    h = ws2["A1"]
    h.value     = f"PF & ESI STATUTORY REGISTER — {month_name.upper()} {year}"
    h.font      = Font(bold=True, size=12, color=BLUE)
    h.alignment = Alignment(horizontal="center", vertical="center")
    h.fill      = _fill(BLUE_LITE)
    ws2.row_dimensions[1].height = 24

    stat_hdrs = ["Sr.","Name","Gross (₹)","Basic (₹)",
                 "PF Employee (₹)","PF Employer (₹)","ESI Employee (₹)","ESI Employer (₹)"]
    stat_widths = [5, 24, 14, 14, 16, 16, 16, 16]
    for i, (h_text, w) in enumerate(zip(stat_hdrs, stat_widths), start=1):
        _hdr(ws2, 2, i, h_text, bg="1A3A5C", wrap=True)
        ws2.column_dimensions[get_column_letter(i)].width = w

    pf_eligible = [e for e in employees if e["pf"]]
    for sr, emp in enumerate(pf_eligible, start=1):
        slip = calc_salary_fn(emp, year, month)
        row  = 2 + sr
        bg   = GRAY if sr % 2 == 0 else "FFFFFF"
        pf_emp  = slip["pf"]
        pf_er   = round(min(slip["basic"], 15000) * 0.12)   # employer same as employee
        esi_emp = slip["esi"] if emp["esi"] else 0
        esi_er  = round(slip["gross"] * 0.0325) if emp["esi"] else 0  # employer 3.25%

        for col, (val, fmt, aln) in enumerate(zip(
            [sr, emp["name"], slip["gross"], slip["basic"],
             pf_emp, pf_er, esi_emp, esi_er],
            [None, None, INR, INR, INR, INR, INR, INR],
            ["center","left","right","right","right","right","right","right"]
        ), start=1):
            _cell(ws2, row, col, val, fmt=fmt, bg=bg, align=aln)

    # ── Sheet 3: Department Summary ───────────────────────────────────────
    ws3 = wb.create_sheet("Dept Summary")
    ws3.merge_cells("A1:F1")
    h3 = ws3["A1"]
    h3.value     = f"DEPARTMENT COST SUMMARY — {month_name.upper()} {year}"
    h3.font      = Font(bold=True, size=12, color=BLUE)
    h3.alignment = Alignment(horizontal="center", vertical="center")
    h3.fill      = _fill(BLUE_LITE)
    ws3.row_dimensions[1].height = 24

    dept_hdrs = ["Department","Headcount","Gross Payroll (₹)",
                 "Total Deductions (₹)","Net Payroll (₹)","Avg Salary (₹)"]
    dept_widths = [22, 12, 18, 18, 16, 16]
    for i, (h_text, w) in enumerate(zip(dept_hdrs, dept_widths), start=1):
        _hdr(ws3, 2, i, h_text, bg="1A3A5C")
        ws3.column_dimensions[get_column_letter(i)].width = w

    dept_map = {}
    for emp in employees:
        slip = calc_salary_fn(emp, year, month)
        d = emp["dept"]
        if d not in dept_map:
            dept_map[d] = {"count":0, "gross":0, "ded":0, "net":0}
        dept_map[d]["count"] += 1
        dept_map[d]["gross"] += slip["gross"]
        dept_map[d]["ded"]   += slip["total_ded"]
        dept_map[d]["net"]   += slip["net"]

    for sr, (dept, d) in enumerate(sorted(dept_map.items()), start=1):
        row = 2 + sr
        bg  = GRAY if sr % 2 == 0 else "FFFFFF"
        avg = d["gross"] / d["count"]
        for col, (val, fmt, aln) in enumerate(zip(
            [dept, d["count"], d["gross"], d["ded"], d["net"], avg],
            [None, None, INR, INR, INR, INR],
            ["left","center","right","right","right","right"]
        ), start=1):
            _cell(ws3, row, col, val, fmt=fmt, bg=bg, align=aln)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf
